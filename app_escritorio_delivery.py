import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from openpyxl import Workbook
from openpyxl import load_workbook
import os
import requests
 
#######################################################
 
def guardar(orden):
	ws = wb.active
	ws.append(orden)
	wb.save(filename = "pedidos.xlsx" )
	f = open("datostarde.csv","a")
	f.write(orden[0]+","+str(orden[1])+","+str(orden[2])+","+str(orden[3])+","+str(orden[4])+"\n")
	f.close()
 

def validar(dato):
	if dato:
		try:
			dato = int(dato)
			return dato
		except ValueError:
			return -1
	else:
		return -1
 
 
def borrar():
	cpedido.delete(0,tk.END)
	ctotal.delete(0,tk.END)
	ccliente.delete(0,tk.END)
	cdireccion.delete(0,tk.END)
	ccomentarios.delete(0,tk.END)
 
 
def pedido():
	pedido = cpedido.get()
	total = ctotal.get()
	direccion = cdireccion.get()
	cliente = ccliente.get()
	comentarios = ccomentarios.get()
	if cliente:
		respuesta = messagebox.askyesno(title="Pregunta", message="¿Confirma el pedido?")
		if respuesta:
			gustos = [cliente,pedido,total,direccion,comentarios]
			#ACCION DE GUARDAR
			guardar(gustos)
			messagebox.showinfo(title="Información", message="Pedido Exitoso")
			borrar()
		else:
			messagebox.showinfo(title="Información", message="Pedido en pausa")
	else:
		messagebox.showwarning(title="Advertencia", message="Debe llenar los recuadros, no sea imbecil")
 

def info():
	messagebox.showinfo(title="Información", message="App 2.0 by Mariano Avola")


def cancelar_pedido():
	respuesta = messagebox.askyesno(title="Pregunta", message="¿Desea cancelar el pedido?")
	if respuesta:
		cpedido.delete(0,tk.END)
		ctotal.delete(0,tk.END)
		ccliente.delete(0,tk.END)
		cdireccion.delete(0,tk.END)
		ccomentarios.delete(0,tk.END)
 

def comprobarArchivo():
	existe = os.path.exists("pedidos.xlsx")
	if existe:
		wb = load_workbook(filename = "pedidos.xlsx")
		ws = wb.active
	else:
		wb = Workbook()
		ws = wb.active
		titulo = ["Nombre","Pedido","Total","Direccion","Comentarios"]
		ws.append(titulo)
		wb.save(filename = "pedidos.xlsx" )
		f = open("datostarde.csv","a")
		f.write("Nombre,Pedido,Total,Direccion,Comentarios\n")
		f.close()
	return wb


#############################################################
 
wb = comprobarArchivo()
 
#############################################################
 

ventana = tk.Tk()
ventana.config(width = 500, height = 500)
ventana.title("Pedidos")
 
 
########## ETIQUETAS ##########
ebienvenido = tk.Label(text="INFORMACION DE PEDIDOS")
ebienvenido.place(x = 170, y = 25)
epedido = tk.Label(text = "Ingrese Pedido : ")
epedido.place(x = 50, y = 90)
etotal = tk.Label(text = "Total : ")
etotal.place(x = 50, y = 140)
ecliente = tk.Label(text = "Nombre del Cliente : ")
ecliente.place(x = 50, y = 190)
edireccion = tk.Label(text = "Direccion : ")
edireccion.place(x = 50, y = 240)
ecomentarios = tk.Label(text = "Comentarios : ")
ecomentarios.place(x = 50, y = 290) 

 
########## CAJAS ##########
cpedido = ttk.Entry()
cpedido.place(x = 200, y = 90, width=250)
ctotal = ttk.Entry()
ctotal.place(x = 200, y = 140, width=250)
ccliente = ttk.Entry()
ccliente.place(x = 200, y = 190, width=250)
cdireccion = ttk.Entry()
cdireccion.place(x = 200, y = 240, width=250)
ccomentarios = ttk.Entry()
ccomentarios.place(x = 200, y = 290, width=250)

 
########## BOTONES ##########
bpedido = ttk.Button(text = "Hacer Pedido", command = pedido)
bpedido.place(x = 350 , y = 430, height=40, width = 100)
 
bcancelar = ttk.Button(text = "Cancelar Pedido", command = cancelar_pedido)
bcancelar.place(x = 200 , y = 430, height=40, width = 100)
 
binfo = ttk.Button(text = "Info", command = info)
binfo.place(x = 50 , y = 430, height=40, width = 100)
 
 
ventana.mainloop()